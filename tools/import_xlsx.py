"""One-shot importer: 'Gestion Salaire .xlsx' -> GestionMoney SQLite.

Reads every monthly sheet, rebuilds the fixed-expense catalogue, the per-month
envelopes with their planned/paid figures, every one-off spend, the account
balance snapshots and the salary line.

Usage:  python tools/import_xlsx.py "C:\\path\\to\\Gestion Salaire .xlsx" [--wipe]
"""
import os
import re
import sys
import calendar

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import db
import api

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is needed for the import only:\n    pip install openpyxl")

MONTHS = {
    "janvier": 1, "january": 1, "fevrier": 2, "february": 2, "mars": 3, "march": 3,
    "avril": 4, "april": 4, "mai": 5, "may": 5, "juin": 6, "june": 6,
    "juillet": 7, "july": 7, "aout": 8, "august": 8, "septembre": 9, "september": 9,
    "spetember": 9, "octobre": 10, "october": 10, "novembre": 11, "november": 11,
    "decembre": 12, "december": 12,
}

# label (lowercased, trimmed) -> (canonical label, category, kind)
CATALOG = {
    "loyer":               ("Loyer", "Housing", "bill"),
    "darna":               ("Darna", "Housing", "bill"),
    "engie":               ("Engie", "Utilities", "bill"),
    "fibre optique":       ("Fibre Optique", "Utilities", "bill"),
    "forfait mobile":      ("Forfait Mobile", "Utilities", "bill"),
    "pass navigo":         ("Pass Navigo", "Transport", "bill"),
    "credit voiture":      ("Credit Voiture", "Transport", "bill"),
    "repas dehors":        ("Repas Dehors", "Food", "budget"),
    "super marche":        ("Super Marche", "Groceries", "budget"),
    "aide courses maison": ("Aide Courses Maison", "Family", "budget"),
    "free money":          ("Free Money", "Personal", "budget"),
    "unpredicted money":   ("Unpredicted Money", "Other", "budget"),
    "fitness park":        ("Fitness Park", "Health", "bill"),
    "canal plus":          ("Canal Plus", "Subscription", "bill"),
    "livret a":            ("Livret A", "Savings", "saving"),
    "savings":             ("Savings", "Savings", "saving"),
    "bills":               ("Bills", "Utilities", "bill"),
    "removing kyst":       ("Removing Kyst", "Health", "bill"),
}

# keyword -> category, for the free-text one-off spends
KEYWORDS = [
    (r"amazon|aliexpress|ali express|shein|temu|cdiscount|fnac", "Personal"),
    (r"nintendo|nintando|xbox|playstation|steam|jeux|game|microsoft", "Personal"),
    (r"pharmac|medic|docteur|dentist|opticien|lunette|tensionmeter|sante", "Health"),
    (r"cado|cadeau|9adhey|gato|anniversaire|fleur|flower|gift", "Gifts"),
    (r"resto|restau|repas|mcdo|burger|pizza|kfc|starbucks|cafe", "Food"),
    (r"carrefour|lidl|auchan|monoprix|leclerc|course|marche", "Groceries"),
    (r"navigo|uber|velib|taxi|train|sncf|essence|peage|transavia|vol |avion|flight", "Transport"),
    (r"zara|celio|nike|sbedri|pull|shirt|serwel|sirwel|jean|chaussure|decathlon|sephora|coiffeur", "Personal"),
    (r"ikea|leroy merlin|meuble|canap|tawla|farch|frigidaire|machine a laver|tv\b|four", "Housing"),
    (r"netflix|spotify|canal|abonnement|subscription|pass\b|fiver|website|domain", "Subscription"),
    (r"maman|baba|papa|sabta|3ammi|khalti|famille|family|zouhour|latifa", "Family"),
    (r"livret|epargne|invest|saving", "Savings"),
]

BAL_MAP = {
    "current livret a": "Livret A",
    "cuurent sg": "Compte",
    "current sg": "Compte",
    "current rivolute": "Revolut",
    "current revolute": "Revolut",
    "current cash": "Cash",
}


def norm(s):
    """Lowercase, strip accents and squeeze whitespace."""
    s = str(s).strip().lower()
    for a, b in (("é", "e"), ("è", "e"), ("ê", "e"), ("à", "a"), ("â", "a"),
                 ("î", "i"), ("ï", "i"), ("ô", "o"), ("û", "u"), ("ù", "u"),
                 ("ç", "c"), ("\ufffd", "e")):
        s = s.replace(a, b)
    return re.sub(r"\s+", " ", s)


def num(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ".").replace("\u00a0", "").strip())
    except ValueError:
        return None


def sheet_ym(title):
    low = norm(title)
    m = re.match(r"^([a-z]+)\s*(\d{4})?$", low)
    if not m:
        return None
    mon = MONTHS.get(m.group(1))
    if not mon:
        return None
    year = int(m.group(2)) if m.group(2) else 2023  # the untitled 2023 run
    return "%04d-%02d" % (year, mon)


def categorise(label):
    n = norm(label)
    for pat, cat in KEYWORDS:
        if re.search(pat, n):
            return cat
    return "Other"


def catalog_entry(label):
    n = norm(label)
    if n in CATALOG:
        return CATALOG[n]
    for key, val in CATALOG.items():
        if n.startswith(key) or key.startswith(n):
            return val
    return (str(label).strip(), categorise(label), "bill")


def run(path, wipe=False, profile=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    profile = profile if profile is not None else db.list_profiles()[0]['id']
    con = db.init(profile)

    if wipe:
        for t in ("tx", "envelopes", "months", "recurring", "balances"):
            con.execute("DELETE FROM " + t)
        con.commit()
        print("Cleared existing budget data.")

    # --- pass 1: collect every sheet ---
    sheets = []
    for ws in wb.worksheets:
        ym = sheet_ym(ws.title)
        if ym:
            sheets.append((ym, ws))
    sheets.sort(key=lambda x: x[0])
    print("Found %d monthly sheets: %s -> %s"
          % (len(sheets), sheets[0][0], sheets[-1][0]))

    # --- pass 2: build the recurring catalogue ---
    # An item is switched ON only if it still appeared in the last 3 months of
    # data, so lines you dropped long ago (Fitness Park, Savings...) stay in the
    # catalogue but do not clutter new months.
    recent_yms = set(ym for ym, _ in sheets[-3:])
    seen = {}
    for ym, ws in sheets:
        for r in range(5, 40):
            a, b = ws.cell(r, 1).value, num(ws.cell(r, 2).value)
            if not (isinstance(a, str) and a.strip() and b is not None):
                continue
            if "chaque" in norm(a):
                continue
            lab, cat, kind = catalog_entry(a)
            e = seen.setdefault(lab, {"cat": cat, "kind": kind, "n": 0,
                                      "last": 0.0, "recent": False})
            e["n"] += 1
            e["last"] = b
            if ym in recent_yms:
                e["recent"] = True

    order = sorted(seen.items(), key=lambda kv: -kv[1]["n"])
    rec_id = {}
    n_active = 0
    for i, (lab, e) in enumerate(order):
        row = api.one(con.execute("SELECT id FROM recurring WHERE label=?", (lab,)))
        active = 1 if e["recent"] else 0
        n_active += active
        if row:
            rec_id[lab] = row["id"]
        else:
            cur = con.execute(
                "INSERT INTO recurring(label,category,kind,amount,active,sort)"
                " VALUES(?,?,?,?,?,?)",
                (lab, e["cat"], e["kind"], round(e["last"], 2), active, i * 10))
            rec_id[lab] = cur.lastrowid
    con.commit()
    print("Recurring catalogue: %d items (%d active)." % (len(order), n_active))

    n_env = n_tx = n_bal = 0

    # --- pass 3: months, envelopes, transactions, balances ---
    for ym, ws in sheets:
        y, mo = int(ym[:4]), int(ym[5:7])
        last_day = calendar.monthrange(y, mo)[1]
        mid_day = min(15, last_day)

        salary = num(ws.cell(1, 2).value) or 0.0
        extra = num(ws.cell(1, 3).value) or 0.0
        if extra and extra > salary:      # C1 sometimes held the gross, not a bonus
            extra = round(extra - salary, 2)

        con.execute("INSERT INTO months(ym,income,extra_income,note,closed)"
                    " VALUES(?,?,?,?,1)"
                    " ON CONFLICT(ym) DO UPDATE SET income=excluded.income,"
                    " extra_income=excluded.extra_income, closed=1",
                    (ym, salary, extra, "imported from " + ws.title))
        m = api.one(con.execute("SELECT * FROM months WHERE ym=?", (ym,)))
        mid = m["id"]

        # fixed / budget lines -> envelopes (+ one tx carrying what was actually paid)
        for r in range(5, 40):
            a, b = ws.cell(r, 1).value, num(ws.cell(r, 2).value)
            if not (isinstance(a, str) and a.strip() and b is not None):
                continue
            if "chaque" in norm(a):
                continue
            lab, cat, kind = catalog_entry(a)
            paid = num(ws.cell(r, 3).value) or 0.0
            con.execute(
                "INSERT OR IGNORE INTO envelopes"
                " (month_id,recurring_id,label,category,kind,planned,sort)"
                " VALUES(?,?,?,?,?,?,?)",
                (mid, rec_id.get(lab), lab, cat, kind, round(b, 2), r))
            env = api.one(con.execute(
                "SELECT id FROM envelopes WHERE month_id=? AND label=?", (mid, lab)))
            n_env += 1
            if paid > 0 and env:
                con.execute(
                    "INSERT INTO tx(date,label,amount,kind,envelope_id,category,"
                    " account,note,oneoff) VALUES(?,?,?,?,?,?,?,?,0)",
                    ("%s-%02d" % (ym, mid_day), lab, round(paid, 2), "expense",
                     env["id"], cat, "Compte",
                     "imported total from xlsx - split it into real transactions"))
                n_tx += 1

        # one-off spends -> transactions flagged oneoff
        for r in range(5, 40):
            f, g = ws.cell(r, 6).value, num(ws.cell(r, 7).value)
            if not (isinstance(f, str) and f.strip() and g is not None):
                continue
            nf = norm(f)
            if nf.startswith(("current", "cuurent", "a payer", "old ", "free money",
                              "eom", "new livret")):
                continue
            con.execute(
                "INSERT INTO tx(date,label,amount,kind,category,account,note,oneoff)"
                " VALUES(?,?,?,'expense',?,?,?,1)",
                ("%s-%02d" % (ym, mid_day), f.strip(), round(g, 2),
                 categorise(f), "Compte", "imported from " + ws.title))
            n_tx += 1

        # account balances -> snapshots
        for r in range(30, 50):
            for c in (3, 6, 9):
                lbl = ws.cell(r, c).value
                val = num(ws.cell(r, c + 1).value)
                if isinstance(lbl, str) and val is not None:
                    acct = BAL_MAP.get(norm(lbl))
                    if acct:
                        con.execute(
                            "INSERT INTO balances(account,ym,amount) VALUES(?,?,?)"
                            " ON CONFLICT(account,ym) DO UPDATE SET amount=excluded.amount",
                            (acct, ym, round(val, 2)))
                        n_bal += 1
        con.commit()

    print("Imported: %d months, %d envelopes, %d transactions, %d balance snapshots."
          % (len(sheets), n_env, n_tx, n_bal))

    # --- shopping list -> inactive recurring "wishlist" is not wanted; skip ---
    tot = api.one(con.execute("SELECT COUNT(*) n, SUM(amount) s FROM tx"))
    print("Database now holds %d transactions worth %.2f."
          % (tot["n"], tot["s"] or 0))
    print("\nDone. Start the app and open the Transactions tab to review.")


if __name__ == "__main__":
    if "--list-profiles" in sys.argv:
        for p in db.list_profiles():
            print("  id=%-3d %s" % (p["id"], p["name"]))
        sys.exit(0)
    prof = None
    if "--profile" in sys.argv:
        prof = int(sys.argv[sys.argv.index("--profile") + 1])
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    args = [a for a in args if a != str(prof)]
    src = args[0] if args else os.path.expanduser("~/Downloads/Gestion Salaire .xlsx")
    if not os.path.isfile(src):
        sys.exit("Cannot find the workbook: " + src)
    run(src, wipe="--wipe" in sys.argv, profile=prof)
